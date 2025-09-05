"""
module containing functions for plotting results from a given results file
and functions for plotting results after creating the corresponding results
"""

from __future__ import annotations
import os
import re
import json
import math
import bz2
import colorsys
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt


# ---------------------------- shared helpers ----------------------------

def _norm(s: str) -> str:
    return re.sub(r'[^a-z0-9]+', '', str(s).strip().lower())

def pick_col(df: pd.DataFrame, names: Iterable[str], contains: bool = False) -> Optional[str]:
    if isinstance(names, str):
        names = [names]
    cmap = {_norm(c): c for c in df.columns}
    for n in names:
        n2 = _norm(n)
        if n2 in cmap:
            return cmap[n2]
    if contains:
        for k, v in cmap.items():
            if any(_norm(n) in k for n in names):
                return v
    return None

def ensure_dir(path: str) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)

def _argsort_by(xs: Sequence[int]) -> List[int]:
    return sorted(range(len(xs)), key=lambda i: xs[i])


# ---------------------------- figure 1 ----------------------------

def plot_figure1(
    folder_path: Optional[str] = None,
    output_path: str = "./own_results/tables and figures/figure1.png",
    show_legend: bool = True
) -> str:
    """
    Plot utilization standard deviation (%) against closure levels for series O and NO.
    Reads Result_(O|NO)_*.xlsx, sheet index 3, takes "Variance"→sqrt→%,
    injects O_0 into NO if missing, sorts by level, and saves a PNG.
    """
    folder_path = folder_path or os.path.join(".", "own_results", "All Assignment")
    data = {'O': {'x': [], 'std': []}, 'NO': {'x': [], 'std': []}}
    o0_std = None

    for filename in os.listdir(folder_path):
        if not filename.endswith(".xlsx"):
            continue
        m = re.match(r"Result_(O|NO)_(\d+)\.xlsx", filename)
        if not m:
            continue
        group = m.group(1)
        index = int(m.group(2))
        file_path = os.path.join(folder_path, filename)
        try:
            df = pd.read_excel(file_path, sheet_name=3)
            if df.shape[1] == 0:
                continue
            df.set_index(df.columns[0], inplace=True)
            variance = df.loc["Variance", df.columns[0]]
            std_dev = float(np.sqrt(float(variance)))
            data[group]['x'].append(index)
            data[group]['std'].append(std_dev)
            if group == 'O' and index == 0:
                o0_std = std_dev
        except Exception:
            continue

    if o0_std is not None and 0 not in data['NO']['x']:
        data['NO']['x'].append(0)
        data['NO']['std'].append(o0_std)

    for group in ['O', 'NO']:
        if not data[group]['x']:
            continue
        order = _argsort_by(data[group]['x'])
        data[group]['x'] = [data[group]['x'][i] for i in order]
        data[group]['std'] = [data[group]['std'][i] for i in order]

    data['O']['std'] = [val * 100 for val in data['O']['std']]
    data['NO']['std'] = [val * 100 for val in data['NO']['std']]

    plt.figure(figsize=(10, 6))
    if data['O']['x']:
        plt.plot(data['O']['x'], data['O']['std'], label='M - Std Dev', linestyle='-', marker='o')
    if data['NO']['x']:
        plt.plot(data['NO']['x'], data['NO']['std'], label='P - Std Dev', linestyle='--', marker='x')
    plt.xlabel("Closure levels")
    plt.ylabel("Utilization Std Deviation (%)")
    if show_legend:
        plt.legend()
    plt.grid(True)
    plt.tight_layout()
    ensure_dir(output_path)
    plt.savefig(output_path)
    plt.close()
    return output_path


# ---------------------------- figure 2 ----------------------------

UK_LAT_MIN, UK_LAT_MAX = 49.0, 61.5
UK_LON_MIN, UK_LON_MAX = -8.8, 2.5

def _normalize_latlon(lat, lon) -> Tuple[Optional[float], Optional[float], bool]:
    try:
        lat = float(lat); lon = float(lon)
    except Exception:
        return None, None, False
    in_uk = (UK_LAT_MIN <= lat <= UK_LAT_MAX) and (UK_LON_MIN <= lon <= UK_LON_MAX)
    if in_uk:
        return lat, lon, False
    swapped_in_uk = (UK_LAT_MIN <= lon <= UK_LAT_MAX) and (UK_LON_MIN <= lat <= UK_LAT_MAX)
    if swapped_in_uk:
        return lon, lat, True
    return lat, lon, False

def _gen_colors(n: int) -> List[str]:
    colors = []
    for i in range(max(1, n)):
        h = (i / max(1, n)) % 1.0
        l = 0.52
        s = 0.70
        r, g, b = colorsys.hls_to_rgb(h, l, s)
        colors.append('#{:02x}{:02x}{:02x}'.format(int(r*255), int(g*255), int(b*255)))
    return colors

def _extract_area(sector_code: str) -> str:
    if not sector_code:
        return ""
    m = re.match(r'^([A-Z]{1,3})', str(sector_code).upper().strip())
    return m.group(1) if m else ""

def plot_figure2(
    sector_xlsx: Optional[str] = None,
    geojson_path: Optional[str] = None,
    boundary_path: Optional[str] = None,
    result_o9: Optional[str] = None,
    result_no9: Optional[str] = None,
    out_dir: str = "./own_results/tables and figures",
    use_basemap: bool = True,
    dpi: int = 200
) -> Tuple[str, str]:
    """
    Create two static PNG maps for O_9 and NO_9. Sectors present in the sector dataset
    are colored by postcode area; open facilities from result files are plotted as points.
    Bounds and styling are kept consistent between the two outputs.
    """
    try:
        import geopandas as gpd
        from shapely.geometry import Point
    except Exception as e:
        raise ImportError("plot_figure2 requires geopandas and shapely.") from e

    try:
        import contextily as cx
        HAS_CTX = True
    except Exception:
        HAS_CTX = False
    if use_basemap and not HAS_CTX:
        use_basemap = False

    sector_xlsx = sector_xlsx or "./data/Sector dataset.xlsx"
    geojson_path = geojson_path or "./data/map_data/all_sectors.geojson"
    boundary_path = boundary_path or "./data/Counties_and_Unitary_Authorities_April_2019_Ultra_Generalised_Boundaries_EW_2022_1790725369940793023.geojson"
    result_o9 = result_o9 or "./own_results/All Assignment/Result_O_9.xlsx"
    result_no9 = result_no9 or "./own_results/All Assignment/Result_NO_9.xlsx"

    sec = pd.read_excel(sector_xlsx, sheet_name=0)
    sector_col = pick_col(sec, ['Postcode Sectors','postcode sectors','sector','pcdsector','pcd'], contains=True)
    if sector_col is None:
        raise ValueError("Sector column not found in sector_xlsx.")
    sec_list = sec[sector_col].astype(str).str.upper().str.strip().replace({"NAN": ""})
    sec_list = sec_list[sec_list != ""]
    have_sectors = set(sec_list.dropna().unique().tolist())
    sec_lat_col  = pick_col(sec, ['rc_centroid_lat','lat','latitude'], contains=True)
    sec_lon_col  = pick_col(sec, ['rc_centroid_lon','lon','longitude'], contains=True)
    sec_name_col = pick_col(sec, ['facility_name','Facility_name','name'], contains=True)

    with open(geojson_path, "r", encoding="utf-8") as f:
        gj = json.load(f)

    present_feats, absent_feats = [], []
    for feat in gj.get("features", []):
        props = feat.get("properties", {})
        sec_code = str(props.get("sector", "")).upper().strip()
        props["sector"] = sec_code
        props["area"]   = _extract_area(sec_code)
        feat["properties"] = props
        (present_feats if sec_code in have_sectors else absent_feats).append(feat)

    gdf_present = gpd.GeoDataFrame.from_features(present_feats, crs="EPSG:4326") if present_feats else gpd.GeoDataFrame(geometry=[])
    gdf_absent  = gpd.GeoDataFrame.from_features(absent_feats,  crs="EPSG:4326") if absent_feats  else gpd.GeoDataFrame(geometry=[])

    areas = sorted(a for a in (gdf_present["area"].fillna("").unique().tolist() if not gdf_present.empty else []) if a)
    palette = _gen_colors(len(areas)) if areas else []
    area_colors = {a: palette[i] for i, a in enumerate(areas)} if areas else {}

    def build_points_from_result(xlsx_path: str) -> List[Tuple[float, float, str]]:
        try:
            u = pd.read_excel(xlsx_path, sheet_name="Utilization Summary")
        except Exception:
            return []
        fac_col_u  = pick_col(u, ['facility','facility_id','zipcode','pcd','sector'], contains=True)
        name_col_u = pick_col(u, ['facility_name','Facility_name','name'], contains=True)
        if fac_col_u is None:
            return []
        candidates = [c for c in [
            pick_col(sec, ['facility','facility_id','zipcode','pcd','sector','Postcode Sectors','pcdsector'], contains=True),
            sector_col
        ] if c]
        u_vals = set(str(x).upper().strip() for x in u[fac_col_u].dropna().astype(str))
        best_col, best_hit = None, -1
        for c in candidates:
            s_vals = set(str(x).upper().strip() for x in sec[c].dropna().astype(str))
            hit = len(u_vals & s_vals)
            if hit > best_hit:
                best_hit, best_col = hit, c
        if best_col is None or best_hit == 0:
            return []
        if not sec_lat_col or not sec_lon_col:
            return []
        u_small = u[[fac_col_u] + ([name_col_u] if name_col_u else [])].drop_duplicates()
        s_small = sec[[best_col, sec_lat_col, sec_lon_col] + ([sec_name_col] if sec_name_col else [])].drop_duplicates()
        merged = u_small.merge(s_small, left_on=fac_col_u, right_on=best_col, how="left")
        pts = []
        for _, r in merged.iterrows():
            lat, lon = r.get(sec_lat_col), r.get(sec_lon_col)
            lat2, lon2, _ = _normalize_latlon(lat, lon)
            if lat2 is None or lon2 is None or pd.isna(lat2) or pd.isna(lon2):
                continue
            nm = ""
            if name_col_u and pd.notna(r.get(name_col_u)):
                nm = str(r.get(name_col_u))
            elif sec_name_col and pd.notna(r.get(sec_name_col)):
                nm = str(r.get(sec_name_col))
            pts.append((lat2, lon2, nm))
        return pts

    def make_png_map(open_facility_locations, out_png_path, gp_4326, ga_4326, area_colors, areas, boundary_path, use_basemap, dpi):
        try:
            import geopandas as gpd
        except Exception:
            raise
        if use_basemap:
            gp = gp_4326.to_crs(3857) if gp_4326 is not None and not gp_4326.empty else gp_4326
            ga = ga_4326.to_crs(3857) if ga_4326 is not None and not ga_4326.empty else ga_4326
            pts_gdf = None
            if open_facility_locations:
                from shapely.geometry import Point as _Point
                pts_gdf = gpd.GeoDataFrame(
                    geometry=[_Point(lon, lat) for lat, lon, *_ in open_facility_locations],
                    crs="EPSG:4326"
                ).to_crs(3857)
            bdf = None
            if boundary_path and os.path.exists(boundary_path):
                bdf = gpd.read_file(boundary_path)
                if bdf.crs and bdf.crs.to_string() != "EPSG:4326":
                    bdf = bdf.to_crs(epsg=4326)
                name_col = None
                for c in bdf.columns:
                    if str(c).lower() in ("ctyua19nm","ctyua22nm","name","lad19nm"):
                        name_col = c; break
                if name_col is not None:
                    target_names = {"Hampshire", "Southampton", "Portsmouth"}
                    bdf = bdf[bdf[name_col].isin(target_names)]
                if not bdf.empty:
                    bdf = bdf.to_crs(3857)
        else:
            gp = gp_4326
            ga = ga_4326
            pts_gdf = None
            if open_facility_locations:
                from shapely.geometry import Point as _Point
                pts_gdf = gpd.GeoDataFrame(
                    geometry=[_Point(lon, lat) for lat, lon, *_ in open_facility_locations],
                    crs="EPSG:4326"
                )
            bdf = None
            if boundary_path and os.path.exists(boundary_path):
                bdf = gpd.read_file(boundary_path)
                if bdf.crs and bdf.crs.to_string() != "EPSG:4326":
                    bdf = bdf.to_crs(epsg=4326)
                name_col = None
                for c in bdf.columns:
                    if str(c).lower() in ("ctyua19nm","ctyua22nm","name","lad19nm"):
                        name_col = c; break
                if name_col is not None:
                    target_names = {"Hampshire", "Southampton", "Portsmouth"}
                    bdf = bdf[bdf[name_col].isin(target_names)]

        fig, ax = plt.subplots(figsize=(12, 9), dpi=dpi)
        if use_basemap:
            bounds = None
            if bdf is not None and not bdf.empty:
                bounds = bdf.total_bounds
            elif gp is not None and not gp.empty:
                bounds = gp.total_bounds
            elif pts_gdf is not None and not pts_gdf.empty:
                bounds = pts_gdf.total_bounds
            if bounds is not None:
                minx, miny, maxx, maxy = bounds
                pad_x = (maxx - minx) * 0.03
                pad_y = (maxy - miny) * 0.03
                ax.set_xlim(minx - pad_x, maxx + pad_x)
                ax.set_ylim(miny - pad_y, maxy + pad_y)
            try:
                import contextily as cx
                cx.add_basemap(ax, source=cx.providers.CartoDB.PositronNoLabels, attribution=False, zorder=0)
            except Exception:
                ax.set_facecolor("#f0f0f0")
        else:
            ax.set_facecolor("#f0f0f0")

        if bdf is not None and not bdf.empty:
            bdf.boundary.plot(ax=ax, color="black", linewidth=0.8, zorder=10)
        if ga_4326 is not None and not ga_4326.empty:
            (ga_4326.to_crs(3857) if use_basemap else ga_4326).plot(ax=ax, facecolor="#eeeeee", edgecolor="#cccccc", linewidth=0.3, alpha=0.25, zorder=11)
        if gp_4326 is not None and not gp_4326.empty:
            gp2 = gp_4326.copy()
            gp2["__color__"] = gp2["area"].map(lambda a: area_colors.get(a, "#6baed6"))
            (gp2.to_crs(3857) if use_basemap else gp2).plot(ax=ax, color=gp2["__color__"], edgecolor="#444444", linewidth=0.3, alpha=0.7, zorder=12)
        if pts_gdf is not None and not pts_gdf.empty:
            ax.scatter(pts_gdf.geometry.x, pts_gdf.geometry.y, s=40, c="black", marker="o", zorder=20)

        handles = []
        if areas:
            from matplotlib.patches import Patch
            for a in areas:
                handles.append(Patch(facecolor=area_colors.get(a, "#6baed6"), edgecolor="#333333", label=a))
        if handles:
            ncol = 2 if len(handles) <= 16 else 3 if len(handles) <= 30 else 4
            leg = ax.legend(handles=handles, loc="lower right", frameon=True, framealpha=0.9, ncol=ncol, fontsize=8)
            leg.set_zorder(100)

        ax.set_xlabel(""); ax.set_ylabel("")
        ax.set_xticks([]); ax.set_yticks([])
        ensure_dir(out_png_path)
        plt.savefig(out_png_path, bbox_inches="tight")
        plt.close(fig)

    out_a = os.path.join(out_dir, "figure2(A).png")
    out_b = os.path.join(out_dir, "figure2(B).png")
    ensure_dir(out_a); ensure_dir(out_b)
    pts_a = build_points_from_result(result_o9)
    pts_b = build_points_from_result(result_no9)
    make_png_map(pts_a, out_a, gdf_present, gdf_absent, area_colors, areas, boundary_path, use_basemap, dpi)
    make_png_map(pts_b, out_b, gdf_present, gdf_absent, area_colors, areas, boundary_path, use_basemap, dpi)
    return out_a, out_b


# ---------------------------- figure 3 ----------------------------

def _postcode_prefix(s: str) -> str:
    if pd.isna(s): return ""
    m = re.match(r'^([A-Za-z]+)', str(s).strip())
    return m.group(1).upper() if m else ""

def plot_figure3(
    sector_path: Optional[str] = None,
    results_dir: Optional[str] = None,
    output_path: str = "./own_results/tables and figures/figure3.png",
    require_urban: bool = True,
    inland_prefixes: Sequence[str] = ("RG","GU","SP"),
    assign_sheet: str = "Assignments"
) -> str:
    """
    Plot median assignment distance (miles) across closure levels for two user groups:
    Group B = (Urban if available) & postcode prefix in {RG,GU,SP}; Group C = others.
    O_0 is injected into NO_0 for baseline alignment.
    """
    sector_path = sector_path or os.path.join('.', 'data', 'Sector dataset.xlsx')
    results_dir = results_dir or os.path.join('.', 'own_results', 'All Assignment')

    sector_df = pd.read_excel(sector_path)
    zip_col = pick_col(sector_df, ['zipcode','zip','user','id'])
    sector_str_col = pick_col(sector_df, ['postcodesector','postcode sectors','postcode sector','sector','pcd','pcdsector'], contains=True)
    if zip_col is None:
        raise ValueError("Sector dataset must contain 'zipcode'.")

    urban_col = pick_col(
        sector_df,
        ['urban','urban/rural','ruralurban','rural-urban','regional spatial type','settlement type','ruc'],
        contains=True
    )

    zipcodes = pd.to_numeric(sector_df[zip_col], errors='coerce').astype('Int64')
    prefixes = sector_df[sector_str_col].astype(str).map(_postcode_prefix) if sector_str_col is not None else pd.Series([""] * len(sector_df))
    if urban_col is not None:
        urban_raw = sector_df[urban_col].astype(str).str.lower()
        is_urban = urban_raw.str.contains('urban') & ~urban_raw.str.contains('rural only|pure rural')
    else:
        is_urban = pd.Series([True] * len(sector_df))

    mask_pref = prefixes.isin(list(inland_prefixes)) if sector_str_col is not None else pd.Series([True] * len(sector_df))
    mask_urb = is_urban if require_urban else pd.Series([True] * len(sector_df))
    mask_B = mask_pref & mask_urb
    B_ids = set(zipcodes[mask_B].dropna().astype(int).tolist())
    C_ids = set(zipcodes[~mask_B].dropna().astype(int).tolist())

    data = {'O': {'x': [], 'B': [], 'C': []},
            'NO': {'x': [], 'B': [], 'C': []}}
    o0_cache = {'B': None, 'C': None}

    for filename in os.listdir(results_dir):
        if not filename.endswith('.xlsx'):
            continue
        m = re.match(r"Result_(O|NO)_(\d+)\.xlsx", filename)
        if not m:
            continue
        group = m.group(1); idx = int(m.group(2))
        fpath = os.path.join(results_dir, filename)
        try:
            df = pd.read_excel(fpath, sheet_name=assign_sheet)
            user_col = pick_col(df, ['user'])
            dist_col = pick_col(df, ['distance','distance(miles)','distance_miles'], contains=True)
            if user_col is None or dist_col is None:
                continue
            users = pd.to_numeric(df[user_col], errors='coerce').astype('Int64')
            dists = pd.to_numeric(df[dist_col], errors='coerce')
            B_mask = users.isin(list(B_ids))
            C_mask = users.isin(list(C_ids))
            B_med = float(dists[B_mask].median()) if B_mask.sum() else np.nan
            C_med = float(dists[C_mask].median()) if C_mask.sum() else np.nan
            data[group]['x'].append(idx)
            data[group]['B'].append(B_med)
            data[group]['C'].append(C_med)
            if group == 'O' and idx == 0:
                o0_cache['B'] = B_med
                o0_cache['C'] = C_med
        except Exception:
            continue

    if o0_cache['B'] is not None and o0_cache['C'] is not None and 0 not in data['NO']['x']:
        data['NO']['x'].append(0)
        data['NO']['B'].append(o0_cache['B'])
        data['NO']['C'].append(o0_cache['C'])

    for g in ['O','NO']:
        if not data[g]['x']:
            continue
        order = _argsort_by(data[g]['x'])
        for key in ['x','B','C']:
            data[g][key] = [data[g][key][i] for i in order]

    plt.figure(figsize=(12,6))
    if data['O']['x']:
        plt.plot(data['O']['x'], data['O']['B'], marker='o', label='MCP - Group A (median distance)')
        plt.plot(data['O']['x'], data['O']['C'], marker='s', label='MCP - Group B (median distance)')
    if data['NO']['x']:
        plt.plot(data['NO']['x'], data['NO']['B'], linestyle='--', marker='x', label='CCP - Group A (median distance)')
        plt.plot(data['NO']['x'], data['NO']['C'], linestyle='--', marker='d', label='CCP - Group B (median distance)')
    plt.xlabel('Closure levels')
    plt.ylabel('Median assignment distance (miles)')
    plt.grid(True)
    plt.legend()
    plt.tight_layout()
    ensure_dir(output_path)
    plt.savefig(output_path)
    plt.close()
    return output_path


# ---------------------------- figure 4 ----------------------------

def plot_figure4(
    folder_path: Optional[str] = None,
    output_path: str = "./own_results/tables and figures/figure4.png",
    show_legend: bool = True
) -> str:
    """
    Plot total saving (£) against closure levels for O and NO using the 'Saving Summary'
    sheet's 'saving (£)' total. Inject O_0 into NO if missing, sort by level, and save PNG.
    """
    folder_path = folder_path or os.path.join(".", "own_results", "All Assignment")
    data = {'O': {'x': [], 'val': []}, 'NO': {'x': [], 'val': []}}
    o0_value = None

    for filename in os.listdir(folder_path):
        if not filename.endswith(".xlsx"):
            continue
        m = re.match(r"Result_(O|NO)_(\d+)\.xlsx", filename)
        if not m:
            continue
        group = m.group(1)
        idx = int(m.group(2))
        file_path = os.path.join(folder_path, filename)
        try:
            df = pd.read_excel(file_path, sheet_name="Saving Summary")
            if df.columns.size > 0:
                df = df.set_index(df.columns[0])
            df.index = df.index.map(lambda x: str(x).strip().lower())
            norm_cols = {c: str(c).strip() for c in df.columns}
            df = df.rename(columns=norm_cols)
            candidates = [c for c in df.columns if _norm(c) == _norm("saving (£)")]
            if not candidates or "total" not in df.index:
                continue
            saving_col = candidates[0]
            total_value = float(df.loc["total", saving_col])
            if group == 'O' and idx == 0:
                o0_value = total_value
            data[group]['x'].append(idx)
            data[group]['val'].append(total_value)
        except Exception:
            continue

    if o0_value is not None and 0 not in data['NO']['x']:
        data['NO']['x'].append(0)
        data['NO']['val'].append(o0_value)

    for group in ['O', 'NO']:
        if not data[group]['x']:
            continue
        order = _argsort_by(data[group]['x'])
        data[group]['x'] = [data[group]['x'][i] for i in order]
        data[group]['val'] = [data[group]['val'][i] for i in order]

    plt.figure(figsize=(10, 6))
    if data['O']['x']:
        plt.plot(data['O']['x'], data['O']['val'], label="MCP - Saving (£)", marker='o')
    if data['NO']['x']:
        plt.plot(data['NO']['x'], data['NO']['val'], label="CCP - Saving (£)", marker='x', linestyle='--')
    plt.xlabel("Closure levels")
    plt.ylabel("Saving (£ million)")
    if show_legend:
        plt.legend()
    plt.grid(True)
    plt.tight_layout()
    ensure_dir(output_path)
    plt.savefig(output_path)
    plt.close()
    return output_path


# ---------------------------- table 5 ----------------------------

from decimal import Decimal, ROUND_HALF_UP

def _to_int_or_none(x):
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return None
    try:
        return int(Decimal(str(x)).quantize(0, rounding=ROUND_HALF_UP))
    except Exception:
        return None

def _round3_or_none(x):
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return None
    try:
        return float(Decimal(str(x)).quantize(Decimal('0.000'), rounding=ROUND_HALF_UP))
    except Exception:
        return None

def _diff_int(a, b):
    if a is None or b is None:
        return None
    return _to_int_or_none(float(a) - float(b))

def _median_min_nearest_miles(user_coords: np.ndarray, fac_coords: np.ndarray) -> Optional[float]:
    if user_coords.size == 0 or fac_coords.size == 0:
        return None
    R = 3958.7613
    users = np.radians(user_coords)
    facs  = np.radians(fac_coords)
    phi1 = users[:, [0]]; lam1 = users[:, [1]]
    phi2 = facs[:, 0][None, :]; lam2 = facs[:, 1][None, :]
    dphi = phi2 - phi1; dlam = lam2 - lam1
    a = np.sin(dphi / 2) ** 2 + np.cos(phi1) * np.cos(phi2) * np.sin(dlam / 2) ** 2
    d = 2 * R * np.arcsin(np.sqrt(a))
    mins = d.min(axis=1)
    return float(np.median(mins))

def _get_closed_facility_ids(xlsx_path: str) -> set:
    try:
        df = pd.read_excel(xlsx_path, sheet_name='Closed Facilities')
    except Exception:
        return set()
    if df.shape[1] < 3:
        return set()
    mask = df.iloc[:, 0:3].notna().all(axis=1)
    fid_col = pick_col(df, ['facility_id', 'facility id', 'fid'])
    if not fid_col:
        return set()
    M = df.loc[mask, fid_col].dropna()
    return set(M.astype(str).str.strip().tolist())

def _get_affected_user_map(baseline_xlsx_path: str, M: set) -> Dict[str, str]:
    try:
        df = pd.read_excel(baseline_xlsx_path, sheet_name='Assignments')
    except Exception:
        return {}
    fac_col  = pick_col(df, ['facility'])
    user_col = pick_col(df, ['user'])
    if not fac_col or not user_col:
        return {}
    df[fac_col]  = df[fac_col].astype(str).str.strip()
    df[user_col] = df[user_col].astype(str).str.strip()
    affected = df[df[fac_col].isin(M)][[user_col, fac_col]].dropna()
    return dict(zip(affected[user_col], affected[fac_col]))

def _compute_median_distance(sector_df: pd.DataFrame, N: set, M: set,
                             ZIPCOL: str, USER_LAT: str, USER_LON: str, FAC_LAT: str, FAC_LON: str, CAPCOL: str) -> Optional[float]:
    users_df = sector_df[sector_df[ZIPCOL].isin(N)][[USER_LAT, USER_LON]].dropna()
    I = users_df[[USER_LAT, USER_LON]].to_numpy()
    open_df = sector_df[(~sector_df[ZIPCOL].isin(M)) & (sector_df[CAPCOL] > 0)][[FAC_LAT, FAC_LON]].dropna()
    J = open_df[[FAC_LAT, FAC_LON]].to_numpy()
    return _median_min_nearest_miles(I, J)

def _sum_pnj_times_population(travel_dict: Optional[dict], sector_df: pd.DataFrame, user_to_fac: Dict[str, str], ZIPCOL: str, POPCOL: str) -> Optional[float]:
    if travel_dict is None or not user_to_fac:
        return None
    total = 0.0
    for n_str, j_str in user_to_fac.items():
        try:
            n_int = int(n_str); j_int = int(j_str)
        except Exception:
            continue
        p = travel_dict.get(n_int, {}).get(j_int, 0.0)
        pop_rows = sector_df[sector_df[ZIPCOL] == n_str]
        if pop_rows.empty:
            continue
        pop_val = pop_rows[POPCOL].iloc[0]
        if pd.isna(pop_val):
            continue
        total += float(p) * float(pop_val)
    return float(total)

def _load_travel_dict(path: str) -> Optional[dict]:
    if not path or not os.path.exists(path):
        return None
    with bz2.BZ2File(path, 'rb') as f:
        try:
            d = json.load(f) if path.endswith('.json.bz2') else None
        except Exception:
            d = None
    if d is None:
        import pickle as cPickle
        try:
            with bz2.BZ2File(path, 'rb') as f:
                d = cPickle.load(f)
        except Exception:
            return None
    try:
        return {int(i): {int(j): float(d[i][j]) for j in d[i]} for i in d}
    except Exception:
        return None

def build_table5(
    assign_dir: Optional[str] = None,
    tables_dir: str = "./own_results/tables and figures",
    sector_path: Optional[str] = None,
    travel_dict_path: str = "./data/travel_dict.json.pbz2",
    output_filename: str = "table5.xlsx"
) -> str:
    """
    Build the Table 5 Excel summarizing council vs model metrics for closures 1..9:
    expected users, median nearest-open distance, median traveled distance,
    utilization medians, and differences (Council − Model) with proper formats.
    """
    assign_dir = assign_dir or os.path.join('.', 'own_results', 'All Assignment')
    sector_path = sector_path or os.path.join('.', 'data', 'Sector dataset.xlsx')
    ensure_dir(os.path.join(tables_dir, "_dummy"))

    sector_df = pd.read_excel(sector_path)
    ZIPCOL   = pick_col(sector_df, ['zipcode', 'zip_code', 'zip'])
    USER_LAT = pick_col(sector_df, ['centroid_lat', 'user_centroid_lat'])
    USER_LON = pick_col(sector_df, ['centroid_lon', 'user_centroid_lon'])
    FAC_LAT  = pick_col(sector_df, ['rc_centroid_lat', 'facility_centroid_lat', 'rc_lat'])
    FAC_LON  = pick_col(sector_df, ['rc_centroid_lon', 'facility_centroid_lon', 'rc_lon'])
    CAPCOL   = pick_col(sector_df, ['capacity', 'cap'])
    POPCOL   = pick_col(sector_df, ['population', 'pop'])
    for req in [ZIPCOL, USER_LAT, USER_LON, FAC_LAT, FAC_LON, CAPCOL, POPCOL]:
        if not req:
            raise ValueError("Sector dataset is missing required columns.")
    sector_df[ZIPCOL] = sector_df[ZIPCOL].astype(str).str.strip()

    travel_dict = _load_travel_dict(travel_dict_path)

    council_hwrc = [
        "Fair Oak",  "Hayling Island", "Bishops Waltham", "Hartley Wintney","Alresford" ,
        "Marchwood", "Aldershot", "Hedge End", "Bordon"
    ]
    model_hwrc = [
        "Alresford", "Marchwood", "Fair Oak", "Hayling Island", "Bishops Waltham",
        "Petersfield", "Alton", "Casbrook", "Hartley Wintney"
    ]

    rows = []
    for i in range(1, 10):
        no_file = os.path.join(assign_dir, f"Result_NO_{i}.xlsx")
        o_file  = os.path.join(assign_dir, f"Result_O_{i}.xlsx")

        try:
            no_util = pd.read_excel(no_file, sheet_name='Utilization Stats')
            no_util_median_raw = no_util.loc[no_util['Statistic'] == 'Median', 'Utilization'].values[0]
        except Exception:
            no_util_median_raw = None

        try:
            no_dist = pd.read_excel(no_file, sheet_name='Medians Summary')
            no_dist_median_raw = no_dist.loc[
                no_dist['Median distance'] == 'All users: 50th percentile', 'Distance (miles)'
            ].values[0]
        except Exception:
            no_dist_median_raw = None

        try:
            o_util = pd.read_excel(o_file, sheet_name='Utilization Stats')
            o_util_median_raw = o_util.loc[o_util['Statistic'] == 'Median', 'Utilization'].values[0]
        except Exception:
            o_util_median_raw = None

        try:
            o_dist = pd.read_excel(o_file, sheet_name='Medians Summary')
            o_dist_median_raw = o_dist.loc[
                o_dist['Median distance'] == 'All users: 50th percentile', 'Distance (miles)'
            ].values[0]
        except Exception:
            o_dist_median_raw = None

        M_council = _get_closed_facility_ids(no_file)
        baseline_council = os.path.join(assign_dir, "Result_O_0.xlsx") if i == 1 else os.path.join(assign_dir, f"Result_NO_{i-1}.xlsx")
        user2fac_council = _get_affected_user_map(baseline_council, M_council)
        N_council = set(user2fac_council.keys())

        council_dist_nearest_raw = _compute_median_distance(sector_df, N_council, M_council, ZIPCOL, USER_LAT, USER_LON, FAC_LAT, FAC_LON, CAPCOL)
        council_sum_pnj_pop_raw  = _sum_pnj_times_population(travel_dict, sector_df, user2fac_council, ZIPCOL, POPCOL)

        council_sum_pnj_pop  = _to_int_or_none(council_sum_pnj_pop_raw)
        council_dist_nearest = _round3_or_none(council_dist_nearest_raw)
        no_dist_median       = _round3_or_none(no_dist_median_raw)
        no_util_median       = float(no_util_median_raw) if no_util_median_raw is not None else None

        M_model = _get_closed_facility_ids(o_file)
        baseline_model = os.path.join(assign_dir, f"Result_O_{i-1}.xlsx")
        user2fac_model = _get_affected_user_map(baseline_model, M_model)
        N_model = set(user2fac_model.keys())

        model_dist_nearest_raw = _compute_median_distance(sector_df, N_model, M_model, ZIPCOL, USER_LAT, USER_LON, FAC_LAT, FAC_LON, CAPCOL)
        model_sum_pnj_pop_raw  = _sum_pnj_times_population(travel_dict, sector_df, user2fac_model, ZIPCOL, POPCOL)

        model_sum_pnj_pop  = _to_int_or_none(model_sum_pnj_pop_raw)
        model_dist_nearest = _round3_or_none(model_dist_nearest_raw)
        o_dist_median      = _round3_or_none(o_dist_median_raw)
        o_util_median      = float(o_util_median_raw) if o_util_median_raw is not None else None

        diff_exp_users = _diff_int(council_sum_pnj_pop_raw, model_sum_pnj_pop_raw)
        yards_per_mile = 1760.0
        diff_near_open_yards = _to_int_or_none((council_dist_nearest_raw - model_dist_nearest_raw) * yards_per_mile) if (council_dist_nearest_raw is not None and model_dist_nearest_raw is not None) else None
        diff_traveled_yards = _to_int_or_none((no_dist_median_raw - o_dist_median_raw) * yards_per_mile) if (no_dist_median_raw is not None and o_dist_median_raw is not None) else None
        diff_utilization_raw = (float(no_util_median_raw) - float(o_util_median_raw)) if (no_util_median_raw is not None and o_util_median_raw is not None) else None

        rows.append({
            'Closed': i,
            'HWRC (Council)': council_hwrc[i - 1],
            'Baseline expected users of the facility being closed (Council)': council_sum_pnj_pop,
            'Median distance to nearest open fac (miles) (Council)': council_dist_nearest,
            'Median distance traveled (Council)': no_dist_median,
            'Utilization median (Council)': no_util_median,
            'HWRC (Model)': model_hwrc[i - 1],
            'Baseline expected users of the facility being closed (Model)': model_sum_pnj_pop,
            'Median distance to nearest open fac (miles) (Model)': model_dist_nearest,
            'Median distance traveled (Model)': o_dist_median,
            'Utilization median (Model)': o_util_median,
            'Baseline expected users (Diff, Council − Model)': diff_exp_users,
            'Median distance to nearest open fac (yards) (Diff)': diff_near_open_yards,
            'Median distance traveled (yards) (Diff)': diff_traveled_yards,
            'Utilization median (Diff)': diff_utilization_raw,
        })

    df = pd.DataFrame(rows)
    output_path = os.path.join(tables_dir, output_filename)
    df.to_excel(output_path, index=False, startrow=1)

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
        wb = load_workbook(output_path)
        ws = wb.active
        ws.cell(row=1, column=1, value="")
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)
        cell = ws.cell(row=1, column=2, value="Council")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=11)
        cell = ws.cell(row=1, column=7, value="Model")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=1, start_column=12, end_row=1, end_column=15)
        cell = ws.cell(row=1, column=12, value="Difference (Council − Model)")
        cell.alignment = Alignment(horizontal='center', vertical='center')

        for col_idx in [4, 5, 9, 10]:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.000'
        for col_idx in [3, 8, 12, 13, 14]:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0'
        for col_idx in [6, 11, 15]:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.000%'
        wb.save(output_path)
    except Exception:
        pass

    return output_path
